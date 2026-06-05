import os
from datetime import datetime, date, timedelta
from typing import List, Dict, Tuple, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib import rcParams
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image as RLImage
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from src.models.models import (
    Employee, TrainingPlan, Enrollment, ExamRecord,
    LearningRecord, Course, Report
)
from src.models.database import SessionLocal
from src.utils.common import generate_code, BASE_DIR
from src.utils.logger import log_operation, log_info, log_error

rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei']
rcParams['axes.unicode_minus'] = False


class ReportGenerator:
    def __init__(self, db: Optional[Session] = None):
        self.db = db or SessionLocal()
        self.report_dir = os.path.join(BASE_DIR, 'exports', 'reports')
        self.chart_dir = os.path.join(self.report_dir, 'charts')
        os.makedirs(self.report_dir, exist_ok=True)
        os.makedirs(self.chart_dir, exist_ok=True)

    def generate_monthly_report(self, year: int, month: int, operator: str) -> Tuple[Optional[Report], Dict, str]:
        try:
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)

            last_month_start = start_date - timedelta(days=1)
            last_month_start = date(last_month_start.year, last_month_start.month, 1)
            last_month_end = start_date - timedelta(days=1)

            current_data = self._calculate_metrics(start_date, end_date)
            previous_data = self._calculate_metrics(last_month_start, last_month_end)

            comparison = self._calculate_comparison(current_data, previous_data)

            report_code = generate_code('REPORT_')
            period = f"{year}年{month}月"

            report = Report(
                report_code=report_code,
                report_type='monthly',
                period=period,
                start_date=start_date,
                end_date=end_date,
                completion_rate=current_data['completion_rate'],
                average_score=current_data['average_score'],
                participation_rate=current_data['participation_rate'],
                total_training_hours=current_data['total_training_hours'],
                total_participants=current_data['unique_participants'],
                created_by=operator
            )

            self.db.add(report)
            self.db.flush()

            chart_paths = self._generate_charts(current_data, previous_data, year, month)

            pdf_path = self._generate_pdf_report(report, current_data, comparison, chart_paths, year, month)
            excel_path = self._generate_excel_report(report, current_data, comparison, chart_paths, year, month)

            report.pdf_path = pdf_path
            report.excel_path = excel_path
            self.db.commit()

            log_operation(operator, '生成月度报告', 'Report', report.id,
                          f"报告期: {period}, 完成率: {current_data['completion_rate']:.2f}%")

            result = {
                'report': report,
                'current_data': current_data,
                'comparison': comparison,
                'pdf_path': pdf_path,
                'excel_path': excel_path
            }

            return report, result, "月度报告生成成功"

        except Exception as e:
            self.db.rollback()
            log_error("生成月度报告失败", e)
            return None, {}, f"生成失败: {str(e)}"

    def _calculate_metrics(self, start_date: date, end_date: date) -> Dict:
        try:
            plans = self.db.query(TrainingPlan).filter(
                TrainingPlan.start_time >= datetime.combine(start_date, datetime.min.time()),
                TrainingPlan.end_time <= datetime.combine(end_date, datetime.max.time())
            ).all()

            plan_ids = [p.id for p in plans]

            enrollments = self.db.query(Enrollment).filter(
                Enrollment.training_plan_id.in_(plan_ids),
                Enrollment.is_active == True
            ).all() if plan_ids else []

            completed_enrollments = [e for e in enrollments if e.status == 'completed']
            completion_rate = (len(completed_enrollments) / len(enrollments) * 100) if enrollments else 0

            exam_records = self.db.query(ExamRecord).filter(
                ExamRecord.completed_at >= datetime.combine(start_date, datetime.min.time()),
                ExamRecord.completed_at <= datetime.combine(end_date, datetime.max.time()),
                ExamRecord.status == 'completed'
            ).all()

            passed_records = [r for r in exam_records if r.is_passed]
            average_score = (sum(r.score for r in exam_records) / len(exam_records)) if exam_records else 0

            learning_records = self.db.query(LearningRecord).filter(
                LearningRecord.record_time >= datetime.combine(start_date, datetime.min.time()),
                LearningRecord.record_time <= datetime.combine(end_date, datetime.max.time())
            ).all()

            total_hours = sum(r.study_duration_minutes for r in learning_records) / 60.0

            unique_employees = set()
            for e in enrollments:
                unique_employees.add(e.employee_id)

            total_employees = self.db.query(Employee).count()
            participation_rate = (len(unique_employees) / total_employees * 100) if total_employees else 0

            hours_distribution = self._calculate_hours_distribution(start_date, end_date)
            department_stats = self._calculate_department_stats(start_date, end_date)
            course_stats = self._calculate_course_stats(start_date, end_date)

            return {
                'total_training_plans': len(plans),
                'total_enrollments': len(enrollments),
                'completed_enrollments': len(completed_enrollments),
                'completion_rate': round(completion_rate, 2),
                'total_exams': len(exam_records),
                'passed_exams': len(passed_records),
                'pass_rate': round((len(passed_records) / len(exam_records) * 100) if exam_records else 0, 2),
                'average_score': round(average_score, 2),
                'total_training_hours': round(total_hours, 2),
                'unique_participants': len(unique_employees),
                'total_employees': total_employees,
                'participation_rate': round(participation_rate, 2),
                'hours_distribution': hours_distribution,
                'department_stats': department_stats,
                'course_stats': course_stats
            }

        except Exception as e:
            log_error("计算指标失败", e)
            return {}

    def _calculate_hours_distribution(self, start_date: date, end_date: date) -> Dict:
        ranges = {
            '0-2小时': 0,
            '2-5小时': 0,
            '5-10小时': 0,
            '10-20小时': 0,
            '20小时以上': 0
        }

        enrollments = self.db.query(Enrollment).filter(
            Enrollment.is_active == True
        ).all()

        for enroll in enrollments:
            hours = enroll.total_study_hours
            if hours < 2:
                ranges['0-2小时'] += 1
            elif hours < 5:
                ranges['2-5小时'] += 1
            elif hours < 10:
                ranges['5-10小时'] += 1
            elif hours < 20:
                ranges['10-20小时'] += 1
            else:
                ranges['20小时以上'] += 1

        return ranges

    def _calculate_department_stats(self, start_date: date, end_date: date) -> List[Dict]:
        employees = self.db.query(Employee).all()
        dept_map = {}
        for emp in employees:
            dept_map[emp.id] = emp.department

        enrollments = self.db.query(Enrollment).all()

        dept_stats = {}
        for enroll in enrollments:
            dept = dept_map.get(enroll.employee_id, '未知部门')
            if dept not in dept_stats:
                dept_stats[dept] = {
                    'department': dept,
                    'enrollments': 0,
                    'completed': 0,
                    'total_hours': 0,
                    'employees': set()
                }
            dept_stats[dept]['enrollments'] += 1
            dept_stats[dept]['employees'].add(enroll.employee_id)
            dept_stats[dept]['total_hours'] += enroll.total_study_hours
            if enroll.status == 'completed':
                dept_stats[dept]['completed'] += 1

        result = []
        for dept, stats in dept_stats.items():
            result.append({
                'department': dept,
                'enrollment_count': stats['enrollments'],
                'completion_count': stats['completed'],
                'completion_rate': round((stats['completed'] / stats['enrollments'] * 100) if stats['enrollments'] else 0, 2),
                'total_hours': round(stats['total_hours'], 2),
                'employee_count': len(stats['employees'])
            })

        result.sort(key=lambda x: x['completion_rate'], reverse=True)
        return result

    def _calculate_course_stats(self, start_date: date, end_date: date) -> List[Dict]:
        plans = self.db.query(TrainingPlan).filter(
            TrainingPlan.start_time >= datetime.combine(start_date, datetime.min.time()),
            TrainingPlan.start_time <= datetime.combine(end_date, datetime.max.time())
        ).all()

        course_stats = {}
        for plan in plans:
            course = plan.course
            if course.id not in course_stats:
                course_stats[course.id] = {
                    'course_id': course.id,
                    'course_title': course.title,
                    'category': course.category,
                    'instructor': course.instructor,
                    'plans': 0,
                    'enrollments': 0,
                    'completed': 0,
                    'avg_score': 0,
                    'total_scores': 0,
                    'score_count': 0
                }
            course_stats[course.id]['plans'] += 1

            for enroll in plan.enrollments:
                if enroll.is_active:
                    course_stats[course.id]['enrollments'] += 1
                    if enroll.status == 'completed':
                        course_stats[course.id]['completed'] += 1

        for exam in self.db.query(ExamRecord).filter(
            ExamRecord.status == 'completed',
            ExamRecord.completed_at >= datetime.combine(start_date, datetime.min.time()),
            ExamRecord.completed_at <= datetime.combine(end_date, datetime.max.time())
        ).all():
            plan = exam.exam.training_plan
            course_id = plan.course_id
            if course_id in course_stats:
                course_stats[course_id]['total_scores'] += exam.score
                course_stats[course_id]['score_count'] += 1

        result = []
        for cid, stats in course_stats.items():
            if stats['score_count'] > 0:
                stats['avg_score'] = round(stats['total_scores'] / stats['score_count'], 2)
            del stats['total_scores']
            del stats['score_count']
            stats['completion_rate'] = round((stats['completed'] / stats['enrollments'] * 100) if stats['enrollments'] else 0, 2)
            result.append(stats)

        result.sort(key=lambda x: x['completion_rate'], reverse=True)
        return result

    def _calculate_comparison(self, current: Dict, previous: Dict) -> Dict:
        def calc_change(curr, prev):
            if prev == 0:
                return None if curr == 0 else 100.0
            return round(((curr - prev) / prev) * 100, 2)

        return {
            'completion_rate_change': calc_change(current.get('completion_rate', 0), previous.get('completion_rate', 0)),
            'average_score_change': calc_change(current.get('average_score', 0), previous.get('average_score', 0)),
            'participation_rate_change': calc_change(current.get('participation_rate', 0), previous.get('participation_rate', 0)),
            'total_training_hours_change': calc_change(current.get('total_training_hours', 0), previous.get('total_training_hours', 0)),
            'total_enrollments_change': calc_change(current.get('total_enrollments', 0), previous.get('total_enrollments', 0)),
            'pass_rate_change': calc_change(current.get('pass_rate', 0), previous.get('pass_rate', 0))
        }

    def _generate_charts(self, current_data: Dict, previous_data: Dict, year: int, month: int) -> Dict[str, str]:
        chart_paths = {}
        prefix = f"{year}{month:02d}"

        chart_paths['completion'] = self._create_bar_chart(
            title='培训完成率对比',
            labels=['上月', '本月'],
            values=[previous_data.get('completion_rate', 0), current_data.get('completion_rate', 0)],
            filename=f'{prefix}_completion.png',
            ylabel='完成率 (%)',
            colors=['#FF9999', '#66B2FF']
        )

        chart_paths['score'] = self._create_bar_chart(
            title='平均成绩对比',
            labels=['上月', '本月'],
            values=[previous_data.get('average_score', 0), current_data.get('average_score', 0)],
            filename=f'{prefix}_score.png',
            ylabel='平均分数',
            colors=['#FFCC99', '#99FF99']
        )

        hours_dist = current_data.get('hours_distribution', {})
        chart_paths['hours'] = self._create_pie_chart(
            title='学时分布',
            labels=list(hours_dist.keys()),
            values=list(hours_dist.values()),
            filename=f'{prefix}_hours.png'
        )

        dept_stats = current_data.get('department_stats', [])
        if dept_stats:
            chart_paths['department'] = self._create_horizontal_bar_chart(
                title='各部门完成率',
                labels=[d['department'] for d in dept_stats],
                values=[d['completion_rate'] for d in dept_stats],
                filename=f'{prefix}_department.png',
                xlabel='完成率 (%)'
            )

        chart_paths['trend'] = self._create_trend_chart(
            title='近6个月培训趋势',
            year=year,
            month=month,
            filename=f'{prefix}_trend.png'
        )

        return chart_paths

    def _create_bar_chart(self, title: str, labels: List[str], values: List[float],
                          filename: str, ylabel: str = '', colors: List[str] = None) -> str:
        fig, ax = plt.subplots(figsize=(8, 5))
        bars = ax.bar(labels, values, color=colors or ['#66B2FF', '#99FF99'])
        ax.set_title(title, fontsize=14, fontweight='bold')
        ax.set_ylabel(ylabel)
        ax.set_ylim(0, max(values) * 1.2 if values else 100)

        for bar, value in zip(bars, values):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2., height,
                    f'{value:.2f}%', ha='center', va='bottom', fontweight='bold')

        plt.tight_layout()
        path = os.path.join(self.chart_dir, filename)
        plt.savefig(path, dpi=100, bbox_inches='tight')
        plt.close()
        return path

    def _create_pie_chart(self, title: str, labels: List[str], values: List[float], filename: str) -> str:
        fig, ax = plt.subplots(figsize=(8, 6))
        colors = ['#FF9999', '#66B2FF', '#99FF99', '#FFCC99', '#C2C2F0']
        wedges, texts, autotexts = ax.pie(values, labels=labels, autopct='%1.1f%%',
                                          colors=colors[:len(values)], startangle=90)
        ax.set_title(title, fontsize=14, fontweight='bold')

        for text in texts:
            text.set_fontsize(10)
        for autotext in autotexts:
            autotext.set_fontsize(9)
            autotext.set_fontweight('bold')

        plt.tight_layout()
        path = os.path.join(self.chart_dir, filename)
        plt.savefig(path, dpi=100, bbox_inches='tight')
        plt.close()
        return path

    def _create_horizontal_bar_chart(self, title: str, labels: List[str], values: List[float],
                                     filename: str, xlabel: str = '') -> str:
        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.barh(labels, values, color='#66B2FF')
        ax.set_title(title, fontsize=14, fontweight='bold')
        ax.set_xlabel(xlabel)
        ax.set_xlim(0, 100)

        for bar, value in zip(bars, values):
            width = bar.get_width()
            ax.text(width + 1, bar.get_y() + bar.get_height() / 2.,
                    f'{value:.2f}%', va='center', fontweight='bold')

        plt.tight_layout()
        path = os.path.join(self.chart_dir, filename)
        plt.savefig(path, dpi=100, bbox_inches='tight')
        plt.close()
        return path

    def _create_trend_chart(self, title: str, year: int, month: int, filename: str) -> str:
        months = []
        completion_rates = []
        avg_scores = []

        for i in range(5, -1, -1):
            m = month - i
            y = year
            if m <= 0:
                m += 12
                y -= 1

            start_date = date(y, m, 1)
            if m == 12:
                end_date = date(y + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(y, m + 1, 1) - timedelta(days=1)

            data = self._calculate_metrics(start_date, end_date)
            months.append(f'{y}年{m}月')
            completion_rates.append(data.get('completion_rate', 0))
            avg_scores.append(data.get('average_score', 0))

        fig, ax1 = plt.subplots(figsize=(12, 6))

        color1 = '#1f77b4'
        ax1.set_xlabel('月份')
        ax1.set_ylabel('完成率 (%)', color=color1)
        line1 = ax1.plot(months, completion_rates, color=color1, marker='o',
                         linewidth=2, markersize=8, label='完成率')
        ax1.tick_params(axis='y', labelcolor=color1)
        ax1.set_ylim(0, 100)

        ax2 = ax1.twinx()
        color2 = '#ff7f0e'
        ax2.set_ylabel('平均分数', color=color2)
        line2 = ax2.plot(months, avg_scores, color=color2, marker='s',
                         linewidth=2, markersize=8, label='平均分数')
        ax2.tick_params(axis='y', labelcolor=color2)
        ax2.set_ylim(0, 100)

        lines = line1 + line2
        labels = [l.get_label() for l in lines]
        ax1.legend(lines, labels, loc='upper left')

        ax1.set_title(title, fontsize=14, fontweight='bold')
        plt.tight_layout()

        path = os.path.join(self.chart_dir, filename)
        plt.savefig(path, dpi=100, bbox_inches='tight')
        plt.close()
        return path

    def _generate_pdf_report(self, report: Report, data: Dict, comparison: Dict,
                             chart_paths: Dict, year: int, month: int) -> str:
        filename = f"training_report_{year}{month:02d}.pdf"
        filepath = os.path.join(self.report_dir, filename)

        doc = SimpleDocTemplate(filepath, pagesize=A4,
                               rightMargin=0.5 * inch, leftMargin=0.5 * inch,
                               topMargin=0.5 * inch, bottomMargin=0.5 * inch)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Title'],
            fontSize=24,
            textColor=colors.darkblue,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.grey,
            alignment=TA_CENTER,
            spaceAfter=15
        )
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.darkblue,
            spaceBefore=15,
            spaceAfter=10
        )
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=8
        )

        story = []

        story.append(Paragraph("培训效能月报", title_style))
        story.append(Paragraph(f"{year}年{month}月", subtitle_style))
        story.append(Spacer(1, 0.3 * inch))

        summary_data = [
            ['指标', '本月数值', '环比变化'],
            ['培训计划数', data['total_training_plans'], self._format_change(None)],
            ['总报名人次', data['total_enrollments'], self._format_change(comparison.get('total_enrollments_change'))],
            ['完成人次', data['completed_enrollments'], self._format_change(None)],
            ['完成率', f"{data['completion_rate']:.2f}%", self._format_change(comparison.get('completion_rate_change'))],
            ['参与人数', data['unique_participants'], self._format_change(None)],
            ['参与率', f"{data['participation_rate']:.2f}%", self._format_change(comparison.get('participation_rate_change'))],
            ['总培训学时', f"{data['total_training_hours']:.2f}", self._format_change(comparison.get('total_training_hours_change'))],
            ['考试人次', data['total_exams'], self._format_change(None)],
            ['通过率', f"{data['pass_rate']:.2f}%", self._format_change(comparison.get('pass_rate_change'))],
            ['平均成绩', f"{data['average_score']:.2f}", self._format_change(comparison.get('average_score_change'))]
        ]

        summary_table = Table(summary_data, colWidths=[2 * inch, 1.5 * inch, 1.5 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(summary_table)

        story.append(PageBreak())

        story.append(Paragraph("核心指标对比", section_style))
        if 'completion' in chart_paths:
            story.append(RLImage(chart_paths['completion'], width=5 * inch, height=3 * inch))
            story.append(Spacer(1, 0.2 * inch))
        if 'score' in chart_paths:
            story.append(RLImage(chart_paths['score'], width=5 * inch, height=3 * inch))

        story.append(PageBreak())

        story.append(Paragraph("学时分布", section_style))
        if 'hours' in chart_paths:
            story.append(RLImage(chart_paths['hours'], width=5 * inch, height=4 * inch))

        story.append(PageBreak())

        story.append(Paragraph("各部门统计", section_style))
        if 'department' in chart_paths:
            story.append(RLImage(chart_paths['department'], width=6 * inch, height=4 * inch))

        dept_data = [['部门', '报名人次', '完成人次', '完成率', '总学时', '参与人数']]
        for d in data.get('department_stats', []):
            dept_data.append([
                d['department'],
                d['enrollment_count'],
                d['completion_count'],
                f"{d['completion_rate']:.2f}%",
                f"{d['total_hours']:.2f}",
                d['employee_count']
            ])

        if len(dept_data) > 1:
            dept_table = Table(dept_data, colWidths=[1.5 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch])
            dept_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(Spacer(1, 0.3 * inch))
            story.append(dept_table)

        story.append(PageBreak())

        story.append(Paragraph("培训趋势分析", section_style))
        if 'trend' in chart_paths:
            story.append(RLImage(chart_paths['trend'], width=7 * inch, height=4 * inch))

        story.append(Paragraph("课程统计", section_style))
        course_data = [['课程名称', '培训次数', '报名人次', '完成率', '平均成绩']]
        for c in data.get('course_stats', [])[:10]:
            course_data.append([
                c['course_title'][:20],
                c['plans'],
                c['enrollments'],
                f"{c['completion_rate']:.2f}%",
                f"{c['avg_score']:.2f}" if c['avg_score'] else '-'
            ])

        if len(course_data) > 1:
            course_table = Table(course_data, colWidths=[2.5 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch])
            course_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkred),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            story.append(Spacer(1, 0.3 * inch))
            story.append(course_table)

        doc.build(story)
        return filepath

    def _format_change(self, change: Optional[float]) -> str:
        if change is None:
            return '-'
        if change > 0:
            return f"<font color='green'>↑ {change:.2f}%</font>"
        elif change < 0:
            return f"<font color='red'>↓ {abs(change):.2f}%</font>"
        else:
            return '0%'

    def _generate_excel_report(self, report: Report, data: Dict, comparison: Dict,
                               chart_paths: Dict, year: int, month: int) -> str:
        filename = f"training_report_{year}{month:02d}.xlsx"
        filepath = os.path.join(self.report_dir, filename)

        wb = openpyxl.Workbook()

        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        cell_font = Font(size=10)
        center_align = Alignment(horizontal='center', vertical='center')

        ws1 = wb.active
        ws1.title = '概览'

        ws1.merge_cells('A1:D1')
        ws1['A1'] = f'{year}年{month}月 培训效能月报'
        ws1['A1'].font = Font(size=16, bold=True, color='1F4E78')
        ws1['A1'].alignment = center_align

        ws1['A3'] = '关键指标'
        ws1['A3'].font = Font(bold=True, size=12)

        headers = ['指标', '本月数值', '环比变化', '说明']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        metrics = [
            ('培训计划数', data['total_training_plans'], None, '当月开设计划总数'),
            ('总报名人次', data['total_enrollments'], comparison.get('total_enrollments_change'), '当月所有培训报名人数'),
            ('完成人次', data['completed_enrollments'], None, '当月完成培训的人数'),
            ('完成率', f"{data['completion_rate']:.2f}%", comparison.get('completion_rate_change'), '完成人次/总报名人次'),
            ('参与人数', data['unique_participants'], None, '当月参与培训的独立员工数'),
            ('参与率', f"{data['participation_rate']:.2f}%", comparison.get('participation_rate_change'), '参与人数/总员工数'),
            ('总培训学时', f"{data['total_training_hours']:.2f}", comparison.get('total_training_hours_change'), '当月累计学习小时数'),
            ('考试人次', data['total_exams'], None, '当月参加考试总次数'),
            ('通过率', f"{data['pass_rate']:.2f}%", comparison.get('pass_rate_change'), '通过人数/考试人次'),
            ('平均成绩', f"{data['average_score']:.2f}", comparison.get('average_score_change'), '所有考试的平均分数')
        ]

        for row, (metric, value, change, note) in enumerate(metrics, 5):
            ws1.cell(row=row, column=1, value=metric).font = cell_font
            ws1.cell(row=row, column=2, value=value).font = cell_font
            ws1.cell(row=row, column=2).alignment = center_align

            change_cell = ws1.cell(row=row, column=3)
            if change is None:
                change_cell.value = '-'
            elif change > 0:
                change_cell.value = f'↑ {change:.2f}%'
                change_cell.font = Font(color='006400', bold=True)
            elif change < 0:
                change_cell.value = f'↓ {abs(change):.2f}%'
                change_cell.font = Font(color='FF0000', bold=True)
            else:
                change_cell.value = '0%'
            change_cell.alignment = center_align

            ws1.cell(row=row, column=4, value=note).font = cell_font

        for col in range(1, 5):
            ws1.column_dimensions[chr(64 + col)].width = 20

        ws2 = wb.create_sheet('部门统计')
        dept_headers = ['部门', '报名人次', '完成人次', '完成率', '总学时', '参与人数']
        for col, header in enumerate(dept_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for row, dept in enumerate(data.get('department_stats', []), 2):
            ws2.cell(row=row, column=1, value=dept['department']).font = cell_font
            ws2.cell(row=row, column=2, value=dept['enrollment_count']).font = cell_font
            ws2.cell(row=row, column=3, value=dept['completion_count']).font = cell_font
            ws2.cell(row=row, column=4, value=f"{dept['completion_rate']:.2f}%").font = cell_font
            ws2.cell(row=row, column=5, value=dept['total_hours']).font = cell_font
            ws2.cell(row=row, column=6, value=dept['employee_count']).font = cell_font

            for col in range(1, 7):
                ws2.cell(row=row, column=col).alignment = center_align

        for col in range(1, 7):
            ws2.column_dimensions[chr(64 + col)].width = 15

        if len(data.get('department_stats', [])) > 1:
            chart1 = BarChart()
            chart1.type = 'bar'
            chart1.title = '各部门完成率'
            chart1.y_axis.title = '部门'
            chart1.x_axis.title = '完成率 (%)'
            chart1.style = 10

            data_ref = Reference(ws2, min_col=4, min_row=1, max_row=len(data['department_stats']) + 1)
            cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=len(data['department_stats']) + 1)
            chart1.add_data(data_ref, titles_from_data=True)
            chart1.set_categories(cats_ref)
            chart1.height = 10
            chart1.width = 20
            ws2.add_chart(chart1, 'H1')

        ws3 = wb.create_sheet('课程统计')
        course_headers = ['课程名称', '培训次数', '报名人次', '完成人次', '完成率', '平均成绩', '讲师', '分类']
        for col, header in enumerate(course_headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for row, course in enumerate(data.get('course_stats', []), 2):
            ws3.cell(row=row, column=1, value=course['course_title']).font = cell_font
            ws3.cell(row=row, column=2, value=course['plans']).font = cell_font
            ws3.cell(row=row, column=3, value=course['enrollments']).font = cell_font
            ws3.cell(row=row, column=4, value=course['completed']).font = cell_font
            ws3.cell(row=row, column=5, value=f"{course['completion_rate']:.2f}%").font = cell_font
            ws3.cell(row=row, column=6, value=f"{course['avg_score']:.2f}" if course['avg_score'] else '-').font = cell_font
            ws3.cell(row=row, column=7, value=course['instructor'] or '-').font = cell_font
            ws3.cell(row=row, column=8, value=course['category'] or '-').font = cell_font

            for col in range(1, 9):
                ws3.cell(row=row, column=col).alignment = center_align

        for col in range(1, 9):
            ws3.column_dimensions[chr(64 + col)].width = 18

        ws4 = wb.create_sheet('学时分布')
        dist_data = data.get('hours_distribution', {})
        ws4.cell(row=1, column=1, value='学时区间')
        ws4.cell(row=1, column=2, value='人数')
        ws4['A1'].fill = header_fill
        ws4['B1'].fill = header_fill
        ws4['A1'].font = header_font
        ws4['B1'].font = header_font
        ws4['A1'].alignment = center_align
        ws4['B1'].alignment = center_align

        for row, (range_, count) in enumerate(dist_data.items(), 2):
            ws4.cell(row=row, column=1, value=range_).font = cell_font
            ws4.cell(row=row, column=2, value=count).font = cell_font
            ws4.cell(row=row, column=1).alignment = center_align
            ws4.cell(row=row, column=2).alignment = center_align

        if dist_data:
            pie = PieChart()
            pie.title = '学时分布'
            labels = Reference(ws4, min_col=1, min_row=2, max_row=len(dist_data) + 1)
            data_pie = Reference(ws4, min_col=2, min_row=1, max_row=len(dist_data) + 1)
            pie.add_data(data_pie, titles_from_data=True)
            pie.set_categories(labels)
            pie.height = 10
            pie.width = 10
            ws4.add_chart(pie, 'D1')

        wb.save(filepath)
        return filepath

    def auto_generate_monthly_report(self, operator: str) -> Tuple[Optional[Report], Dict, str]:
        now = datetime.now()
        last_month = now.month - 1
        year = now.year
        if last_month == 0:
            last_month = 12
            year = now.year - 1

        return self.generate_monthly_report(year, last_month, operator)

    def get_report_list(self, report_type: Optional[str] = None) -> List[Dict]:
        try:
            query = self.db.query(Report)
            if report_type:
                query = query.filter(Report.report_type == report_type)

            reports = query.order_by(Report.created_at.desc()).all()

            result = []
            for r in reports:
                result.append({
                    'report_id': r.id,
                    'report_code': r.report_code,
                    'report_type': r.report_type,
                    'period': r.period,
                    'start_date': r.start_date,
                    'end_date': r.end_date,
                    'completion_rate': r.completion_rate,
                    'average_score': r.average_score,
                    'participation_rate': r.participation_rate,
                    'total_training_hours': r.total_training_hours,
                    'pdf_path': r.pdf_path,
                    'excel_path': r.excel_path,
                    'created_by': r.created_by,
                    'created_at': r.created_at
                })

            return result

        except Exception as e:
            log_error("获取报告列表失败", e)
            return []

    def close(self):
        self.db.close()
