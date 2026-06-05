import os
from datetime import datetime, date
from typing import List, Dict, Tuple, Optional
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from src.models.models import (
    Employee, TrainingPlan, Enrollment, ExamRecord,
    LearningRecord, Course, Exam, Certificate, OperationLog
)
from src.models.database import SessionLocal
from src.utils.common import BASE_DIR
from src.utils.logger import log_operation, log_info, log_error


class DataQueryManager:
    def __init__(self, db: Optional[Session] = None):
        self.db = db or SessionLocal()
        self.export_dir = os.path.join(BASE_DIR, 'exports')
        os.makedirs(self.export_dir, exist_ok=True)

    def query_training_records(self,
                               course_name: Optional[str] = None,
                               instructor: Optional[str] = None,
                               department: Optional[str] = None,
                               start_date: Optional[date] = None,
                               end_date: Optional[date] = None,
                               status: Optional[str] = None) -> List[Dict]:
        try:
            query = self.db.query(
                Enrollment,
                Employee,
                TrainingPlan,
                Course
            ).join(
                Employee, Enrollment.employee_id == Employee.id
            ).join(
                TrainingPlan, Enrollment.training_plan_id == TrainingPlan.id
            ).join(
                Course, TrainingPlan.course_id == Course.id
            )

            if course_name:
                query = query.filter(Course.title.like(f'%{course_name}%'))
            if instructor:
                query = query.filter(Course.instructor.like(f'%{instructor}%'))
            if department:
                query = query.filter(Employee.department.like(f'%{department}%'))
            if start_date:
                query = query.filter(TrainingPlan.start_time >= datetime.combine(start_date, datetime.min.time()))
            if end_date:
                query = query.filter(TrainingPlan.end_time <= datetime.combine(end_date, datetime.max.time()))
            if status:
                query = query.filter(Enrollment.status == status)

            results = query.order_by(TrainingPlan.start_time.desc()).all()

            records = []
            for enroll, emp, plan, course in results:
                records.append({
                    'enrollment_id': enroll.id,
                    'employee_code': emp.employee_id,
                    'employee_name': emp.name,
                    'department': emp.department,
                    'position': emp.position,
                    'course_code': course.course_code,
                    'course_title': course.title,
                    'category': course.category,
                    'instructor': course.instructor,
                    'plan_code': plan.plan_code,
                    'training_start': plan.start_time,
                    'training_end': plan.end_time,
                    'location': plan.location,
                    'enrollment_status': enroll.status,
                    'registered_at': enroll.registered_at,
                    'total_study_hours': enroll.total_study_hours,
                    'engagement_status': enroll.engagement_status,
                    'warning_count': enroll.warning_count
                })

            return records

        except Exception as e:
            log_error("查询培训记录失败", e)
            return []

    def query_exam_scores(self,
                          course_name: Optional[str] = None,
                          instructor: Optional[str] = None,
                          department: Optional[str] = None,
                          start_date: Optional[date] = None,
                          end_date: Optional[date] = None,
                          is_passed: Optional[bool] = None) -> List[Dict]:
        try:
            query = self.db.query(
                ExamRecord,
                Employee,
                Exam,
                TrainingPlan,
                Course
            ).join(
                Employee, ExamRecord.employee_id == Employee.id
            ).join(
                Exam, ExamRecord.exam_id == Exam.id
            ).join(
                TrainingPlan, Exam.training_plan_id == TrainingPlan.id
            ).join(
                Course, TrainingPlan.course_id == Course.id
            )

            if course_name:
                query = query.filter(Course.title.like(f'%{course_name}%'))
            if instructor:
                query = query.filter(Course.instructor.like(f'%{instructor}%'))
            if department:
                query = query.filter(Employee.department.like(f'%{department}%'))
            if start_date:
                query = query.filter(ExamRecord.completed_at >= datetime.combine(start_date, datetime.min.time()))
            if end_date:
                query = query.filter(ExamRecord.completed_at <= datetime.combine(end_date, datetime.max.time()))
            if is_passed is not None:
                query = query.filter(ExamRecord.is_passed == is_passed)

            query = query.filter(ExamRecord.status == 'completed')
            results = query.order_by(ExamRecord.completed_at.desc()).all()

            records = []
            for exam_rec, emp, exam, plan, course in results:
                records.append({
                    'exam_record_id': exam_rec.id,
                    'employee_code': emp.employee_id,
                    'employee_name': emp.name,
                    'department': emp.department,
                    'position': emp.position,
                    'course_code': course.course_code,
                    'course_title': course.title,
                    'instructor': course.instructor,
                    'exam_code': exam.exam_code,
                    'exam_title': exam.title,
                    'attempt_number': exam_rec.attempt_number,
                    'score': exam_rec.score,
                    'is_passed': exam_rec.is_passed,
                    'completed_at': exam_rec.completed_at,
                    'feedback': exam_rec.feedback
                })

            return records

        except Exception as e:
            log_error("查询考试成绩失败", e)
            return []

    def query_operation_logs(self,
                             operator: Optional[str] = None,
                             operation: Optional[str] = None,
                             target_type: Optional[str] = None,
                             start_date: Optional[date] = None,
                             end_date: Optional[date] = None) -> List[Dict]:
        try:
            query = self.db.query(OperationLog)

            if operator:
                query = query.filter(OperationLog.operator.like(f'%{operator}%'))
            if operation:
                query = query.filter(OperationLog.operation.like(f'%{operation}%'))
            if target_type:
                query = query.filter(OperationLog.target_type == target_type)
            if start_date:
                query = query.filter(OperationLog.created_at >= datetime.combine(start_date, datetime.min.time()))
            if end_date:
                query = query.filter(OperationLog.created_at <= datetime.combine(end_date, datetime.max.time()))

            results = query.order_by(OperationLog.created_at.desc()).all()

            records = []
            for log in results:
                records.append({
                    'log_id': log.id,
                    'operator': log.operator,
                    'operation': log.operation,
                    'target_type': log.target_type,
                    'target_id': log.target_id,
                    'details': log.details,
                    'ip_address': log.ip_address,
                    'created_at': log.created_at
                })

            return records

        except Exception as e:
            log_error("查询操作日志失败", e)
            return []

    def export_training_records(self, records: List[Dict], operator: str,
                                filename: Optional[str] = None) -> Tuple[Optional[str], str]:
        try:
            if not records:
                return None, "没有数据可导出"

            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'training_records_{timestamp}.xlsx'

            filepath = os.path.join(self.export_dir, filename)

            df = pd.DataFrame(records)

            column_mapping = {
                'enrollment_id': '报名ID',
                'employee_code': '工号',
                'employee_name': '姓名',
                'department': '部门',
                'position': '岗位',
                'course_code': '课程编号',
                'course_title': '课程名称',
                'category': '课程分类',
                'instructor': '讲师',
                'plan_code': '计划编号',
                'training_start': '培训开始时间',
                'training_end': '培训结束时间',
                'location': '培训地点',
                'enrollment_status': '报名状态',
                'registered_at': '报名时间',
                'total_study_hours': '累计学时',
                'engagement_status': '学习状态',
                'warning_count': '预警次数'
            }

            df = df.rename(columns=column_mapping)

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='培训记录', index=False)

                workbook = writer.book
                worksheet = writer.sheets['培训记录']

                header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True, size=11)
                center_align = Alignment(horizontal='center', vertical='center')

                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column].width = adjusted_width

            log_operation(operator, '导出培训记录', None, None,
                          f"导出记录数: {len(records)}, 文件: {filename}")

            return filepath, f"成功导出{len(records)}条培训记录"

        except Exception as e:
            log_error("导出培训记录失败", e)
            return None, f"导出失败: {str(e)}"

    def export_exam_scores(self, records: List[Dict], operator: str,
                           filename: Optional[str] = None) -> Tuple[Optional[str], str]:
        try:
            if not records:
                return None, "没有数据可导出"

            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'exam_scores_{timestamp}.xlsx'

            filepath = os.path.join(self.export_dir, filename)

            df = pd.DataFrame(records)

            column_mapping = {
                'exam_record_id': '考试记录ID',
                'employee_code': '工号',
                'employee_name': '姓名',
                'department': '部门',
                'position': '岗位',
                'course_code': '课程编号',
                'course_title': '课程名称',
                'instructor': '讲师',
                'exam_code': '考试编号',
                'exam_title': '考试名称',
                'attempt_number': '考试次数',
                'score': '分数',
                'is_passed': '是否通过',
                'completed_at': '完成时间',
                'feedback': '备注'
            }

            df = df.rename(columns=column_mapping)
            df['是否通过'] = df['是否通过'].map({True: '是', False: '否'})

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='考试成绩', index=False)

                workbook = writer.book
                worksheet = writer.sheets['考试成绩']

                header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True, size=11)
                center_align = Alignment(horizontal='center', vertical='center')

                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column].width = adjusted_width

                stats_sheet = workbook.create_sheet('统计汇总')

                total_count = len(records)
                passed_count = sum(1 for r in records if r['is_passed'])
                pass_rate = (passed_count / total_count * 100) if total_count > 0 else 0
                avg_score = (sum(r['score'] for r in records) / total_count) if total_count > 0 else 0

                stats_data = [
                    ['统计项', '数值'],
                    ['总记录数', total_count],
                    ['通过人数', passed_count],
                    ['未通过人数', total_count - passed_count],
                    ['通过率', f'{pass_rate:.2f}%'],
                    ['平均分', f'{avg_score:.2f}'],
                    ['最高分', max(r['score'] for r in records) if records else 0],
                    ['最低分', min(r['score'] for r in records) if records else 0]
                ]

                for row, (key, value) in enumerate(stats_data, 1):
                    cell1 = stats_sheet.cell(row=row, column=1, value=key)
                    cell2 = stats_sheet.cell(row=row, column=2, value=value)
                    if row == 1:
                        cell1.fill = header_fill
                        cell2.fill = header_fill
                        cell1.font = header_font
                        cell2.font = header_font
                    cell1.alignment = center_align
                    cell2.alignment = center_align

                stats_sheet.column_dimensions['A'].width = 15
                stats_sheet.column_dimensions['B'].width = 20

            log_operation(operator, '导出考试成绩', None, None,
                          f"导出记录数: {len(records)}, 文件: {filename}")

            return filepath, f"成功导出{len(records)}条考试成绩记录"

        except Exception as e:
            log_error("导出考试成绩失败", e)
            return None, f"导出失败: {str(e)}"

    def export_operation_logs(self, records: List[Dict], operator: str,
                              filename: Optional[str] = None) -> Tuple[Optional[str], str]:
        try:
            if not records:
                return None, "没有数据可导出"

            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'operation_logs_{timestamp}.xlsx'

            filepath = os.path.join(self.export_dir, filename)

            df = pd.DataFrame(records)

            column_mapping = {
                'log_id': '日志ID',
                'operator': '操作人',
                'operation': '操作类型',
                'target_type': '操作对象',
                'target_id': '对象ID',
                'details': '详情',
                'ip_address': 'IP地址',
                'created_at': '操作时间'
            }

            df = df.rename(columns=column_mapping)

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='操作日志', index=False)

                workbook = writer.book
                worksheet = writer.sheets['操作日志']

                header_fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True, size=11)
                center_align = Alignment(horizontal='center', vertical='center')

                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align

                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 40)
                    worksheet.column_dimensions[column].width = adjusted_width

            log_operation(operator, '导出操作日志', None, None,
                          f"导出记录数: {len(records)}, 文件: {filename}")

            return filepath, f"成功导出{len(records)}条操作日志"

        except Exception as e:
            log_error("导出操作日志失败", e)
            return None, f"导出失败: {str(e)}"

    def get_employee_training_summary(self, employee_id: int) -> Dict:
        try:
            employee = self.db.query(Employee).filter(Employee.id == employee_id).first()
            if not employee:
                return {}

            enrollments = self.db.query(Enrollment).filter(
                Enrollment.employee_id == employee_id,
                Enrollment.is_active == True
            ).all()

            exam_records = self.db.query(ExamRecord).filter(
                ExamRecord.employee_id == employee_id,
                ExamRecord.status == 'completed'
            ).all()

            certificates = self.db.query(Certificate).filter(
                Certificate.employee_id == employee_id
            ).all()

            total_training_hours = sum(e.total_study_hours for e in enrollments)
            completed_count = sum(1 for e in enrollments if e.status == 'completed')
            passed_exams = sum(1 for e in exam_records if e.is_passed)

            avg_score = 0
            if exam_records:
                scores = [e.score for e in exam_records]
                avg_score = sum(scores) / len(scores)

            learning_records = self.db.query(LearningRecord).filter(
                LearningRecord.employee_id == employee_id
            ).all()

            total_learn_minutes = sum(r.study_duration_minutes for r in learning_records)

            return {
                'employee_id': employee.id,
                'employee_code': employee.employee_id,
                'name': employee.name,
                'department': employee.department,
                'position': employee.position,
                'skills': [s.strip() for s in (employee.skills or '').split(',') if s.strip()],
                'total_enrollments': len(enrollments),
                'completed_courses': completed_count,
                'in_progress_courses': len(enrollments) - completed_count,
                'total_training_hours': round(total_training_hours, 2),
                'total_study_minutes': total_learn_minutes,
                'total_exams': len(exam_records),
                'passed_exams': passed_exams,
                'avg_score': round(avg_score, 2),
                'total_certificates': len(certificates),
                'active_certificates': sum(1 for c in certificates if c.status == 'active' and
                                           (not c.valid_until or c.valid_until >= date.today()))
            }

        except Exception as e:
            log_error("获取员工培训汇总失败", e)
            return {}

    def get_department_summary(self, department: Optional[str] = None) -> List[Dict]:
        try:
            query = self.db.query(Employee)
            if department:
                query = query.filter(Employee.department == department)

            employees = query.all()

            dept_employees = {}
            for emp in employees:
                dept = emp.department or '未知部门'
                if dept not in dept_employees:
                    dept_employees[dept] = []
                dept_employees[dept].append(emp.id)

            result = []
            for dept, emp_ids in dept_employees.items():
                enrollments = self.db.query(Enrollment).filter(
                    Enrollment.employee_id.in_(emp_ids),
                    Enrollment.is_active == True
                ).all()

                exam_records = self.db.query(ExamRecord).filter(
                    ExamRecord.employee_id.in_(emp_ids),
                    ExamRecord.status == 'completed'
                ).all()

                total_hours = sum(e.total_study_hours for e in enrollments)
                completed = sum(1 for e in enrollments if e.status == 'completed')
                passed = sum(1 for e in exam_records if e.is_passed)
                avg_score = (sum(e.score for e in exam_records) / len(exam_records)) if exam_records else 0

                result.append({
                    'department': dept,
                    'total_employees': len(emp_ids),
                    'total_enrollments': len(enrollments),
                    'completed_courses': completed,
                    'completion_rate': round((completed / len(enrollments) * 100) if enrollments else 0, 2),
                    'total_training_hours': round(total_hours, 2),
                    'avg_hours_per_employee': round((total_hours / len(emp_ids)) if emp_ids else 0, 2),
                    'total_exams': len(exam_records),
                    'passed_exams': passed,
                    'pass_rate': round((passed / len(exam_records) * 100) if exam_records else 0, 2),
                    'avg_score': round(avg_score, 2)
                })

            result.sort(key=lambda x: x['completion_rate'], reverse=True)
            return result

        except Exception as e:
            log_error("获取部门汇总失败", e)
            return []

    def batch_export(self, export_type: str, operator: str,
                     **query_params) -> Tuple[Optional[str], str]:
        try:
            if export_type == 'training':
                records = self.query_training_records(**query_params)
                return self.export_training_records(records, operator)
            elif export_type == 'exam':
                records = self.query_exam_scores(**query_params)
                return self.export_exam_scores(records, operator)
            elif export_type == 'logs':
                records = self.query_operation_logs(**query_params)
                return self.export_operation_logs(records, operator)
            else:
                return None, f"不支持的导出类型: {export_type}"

        except Exception as e:
            log_error("批量导出失败", e)
            return None, f"导出失败: {str(e)}"

    def close(self):
        self.db.close()
