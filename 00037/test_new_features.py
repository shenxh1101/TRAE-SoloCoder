#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
新功能测试脚本 - API数据同步 + Web文件上传
"""
import sys
import os
import warnings
import time
import json
warnings.filterwarnings('ignore')

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from models.database import init_db, SessionLocal
from models.models import Customer, Order, Receivable, FileUploadRecord, APISyncLog
from modules.api_client import MockDataGenerator, run_mock_sync, DataSyncManager
from utils.helpers import format_currency
import traceback


def print_step(step, title):
    print("\n" + "="*60)
    print(f"  步骤 {step}: {title}")
    print("="*60)


def print_result(success, message):
    status = "✅ 通过" if success else "❌ 失败"
    print(f"\n  {status}: {message}")
    return success


def main():
    print("\n" + "="*60)
    print("  新功能测试 - API数据同步 + Web文件上传")
    print("="*60)

    results = []

    try:
        print_step(1, "初始化数据库")
        try:
            init_db()
            db = SessionLocal()
            customer_count = db.query(Customer).count()
            db.close()
            print(f"  数据库初始化成功，现有 {customer_count} 个客户")

            if customer_count == 0:
                print("  创建示例数据...")
                from main import create_sample_data
                create_sample_data()

            results.append(("数据库初始化", True))
            print_result(True, "数据库初始化完成")
        except Exception as e:
            results.append(("数据库初始化", False))
            print_result(False, f"数据库初始化失败: {e}")
            traceback.print_exc()

        print_step(2, "测试模拟数据生成器")
        try:
            generator = MockDataGenerator()

            print("\n  生成CRM交易记录...")
            transactions = generator.generate_mock_crm_transactions(count=5)
            print(f"  生成了 {len(transactions)} 条CRM交易记录")
            if len(transactions) > 0:
                print(f"  示例: {transactions[0]['customer_name']} - {format_currency(transactions[0]['amount'])}")

            print("\n  生成财务应收账款记录...")
            receivables = generator.generate_mock_finance_receivables(count=4)
            print(f"  生成了 {len(receivables)} 条应收账款记录")
            if len(receivables) > 0:
                print(f"  示例: {receivables[0]['customer_name']} - {format_currency(receivables[0]['total_amount'])}")

            print("\n  生成财务付款记录...")
            payments = generator.generate_mock_finance_payments(count=3)
            print(f"  生成了 {len(payments)} 条付款记录")
            if len(payments) > 0:
                print(f"  示例: {payments[0]['customer_name']} - {format_currency(payments[0]['amount'])}")

            generator.close()
            results.append(("模拟数据生成器", True))
            print_result(True, "模拟数据生成器测试通过")
        except Exception as e:
            results.append(("模拟数据生成器", False))
            print_result(False, f"模拟数据生成器测试失败: {e}")
            traceback.print_exc()

        print_step(3, "测试完整模拟数据同步流程")
        try:
            db = SessionLocal()
            before_count = db.query(Order).count()
            before_receivable_count = db.query(Receivable).count()
            db.close()

            print(f"\n  同步前: 订单数={before_count}, 应收款数={before_receivable_count}")

            sync_result = run_mock_sync()

            db = SessionLocal()
            after_count = db.query(Order).count()
            after_receivable_count = db.query(Receivable).count()
            db.close()

            print(f"\n  同步后: 订单数={after_count}, 应收款数={after_receivable_count}")
            print(f"  新增订单: {after_count - before_count}")
            print(f"  新增应收款: {after_receivable_count - before_receivable_count}")

            all_success = all(v.get('success', False) for v in sync_result.values())
            if all_success:
                results.append(("模拟数据同步流程", True))
                print_result(True, f"模拟数据同步成功，新增 {after_count - before_count} 条订单, {after_receivable_count - before_receivable_count} 条应收款")
            else:
                results.append(("模拟数据同步流程", False))
                print_result(False, "部分同步失败")

        except Exception as e:
            results.append(("模拟数据同步流程", False))
            print_result(False, f"模拟数据同步失败: {e}")
            traceback.print_exc()

        print_step(4, "测试API同步日志记录")
        try:
            db = SessionLocal()
            logs = db.query(APISyncLog).order_by(APISyncLog.created_at.desc()).limit(5).all()
            print(f"\n  找到 {len(logs)} 条同步日志")
            for log in logs:
                print(f"  - [{log.created_at.strftime('%Y-%m-%d %H:%M:%S')}] {log.sync_type}/{log.data_source}: {log.status}, {log.records_synced}条")
            db.close()
            results.append(("API同步日志记录", True))
            print_result(True, f"API同步日志记录正常，共 {len(logs)} 条")
        except Exception as e:
            results.append(("API同步日志记录", False))
            print_result(False, f"API同步日志测试失败: {e}")
            traceback.print_exc()

        print_step(5, "测试Flask Web应用（接口可用性）")
        try:
            from webapp import create_app
            app = create_app()
            client = app.test_client()

            print("\n  测试健康检查接口...")
            response = client.get('/api/health')
            assert response.status_code == 200
            health_data = response.get_json()
            print(f"  健康检查: {health_data['status']}")
            assert health_data['success'] == True

            print("\n  测试获取客户列表接口...")
            response = client.get('/api/customers')
            assert response.status_code == 200
            customer_data = response.get_json()
            print(f"  获取到 {customer_data['total']} 个客户")
            assert customer_data['success'] == True
            assert len(customer_data['data']) > 0

            print("\n  测试获取上传记录列表接口...")
            response = client.get('/api/uploads')
            assert response.status_code == 200
            upload_data = response.get_json()
            print(f"  获取到 {upload_data['total']} 条上传记录")
            assert upload_data['success'] == True

            print("\n  测试主页渲染...")
            response = client.get('/')
            assert response.status_code == 200
            print("  主页渲染成功")

            results.append(("Flask Web应用接口", True))
            print_result(True, "Flask Web应用接口测试全部通过")
        except Exception as e:
            results.append(("Flask Web应用接口", False))
            print_result(False, f"Flask Web应用测试失败: {e}")
            traceback.print_exc()

        print_step(6, "测试Web文件上传接口")
        try:
            from webapp import create_app
            app = create_app()
            client = app.test_client()

            db = SessionLocal()
            customer = db.query(Customer).first()
            db.close()

            print(f"\n  使用测试客户: {customer.name}")

            print("\n  1. 测试缺失文件的情况...")
            response = client.post('/api/upload', data={
                'customer_id': str(customer.id),
                'report_period': '2024-Q4',
                'uploader': '测试人员'
            }, content_type='multipart/form-data')
            assert response.status_code == 400
            result = response.get_json()
            print(f"  正确返回错误: {result['error']}")

            print("\n  2. 测试缺失客户ID的情况...")
            response = client.post('/api/upload', data={
                'report_period': '2024-Q4',
                'uploader': '测试人员'
            }, content_type='multipart/form-data')
            assert response.status_code == 400
            result = response.get_json()
            print(f"  正确返回错误: {result['error']}")

            print("\n  3. 创建测试Excel文件并上传...")
            import openpyxl
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '财务报表'
            ws['A1'] = '项目'
            ws['B1'] = '金额(元)'
            ws['A2'] = '总资产'
            ws['B2'] = 5000000
            ws['A3'] = '总负债'
            ws['B3'] = 2000000
            ws['A4'] = '流动资产'
            ws['B4'] = 2500000
            ws['A5'] = '流动负债'
            ws['B5'] = 1000000
            ws['A6'] = '存货'
            ws['B6'] = 500000
            ws['A7'] = '货币资金'
            ws['B7'] = 800000
            ws['A8'] = '经营活动现金流'
            ws['B8'] = 600000
            ws['A9'] = '营业收入'
            ws['B9'] = 8000000
            ws['A10'] = '净利润'
            ws['B10'] = 800000

            file_stream = BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)

            response = client.post('/api/upload', data={
                'customer_id': str(customer.id),
                'report_period': '2024-Q4',
                'uploader': '测试人员',
                'file': (file_stream, 'test_financial.xlsx')
            }, content_type='multipart/form-data')

            print(f"  响应状态码: {response.status_code}")
            result = response.get_json()

            if result.get('success'):
                print(f"  上传成功!")
                print(f"  资产负债率: {result['data']['financial_indicators']['asset_liability_ratio']}")
                print(f"  流动比率: {result['data']['financial_indicators']['current_ratio']}")
                print(f"  财务健康评分: {result['data']['financial_indicators']['financial_health_score']}")

                db = SessionLocal()
                upload_record = db.query(FileUploadRecord).filter(
                    FileUploadRecord.id == result['data']['upload_id']
                ).first()
                db.close()

                if upload_record:
                    print(f"\n  文件上传记录已保存:")
                    print(f"    - 文件名: {upload_record.file_name}")
                    print(f"    - 文件大小: {upload_record.file_size} bytes")
                    print(f"    - 解析状态: {upload_record.parse_status}")
                    print(f"    - 财务记录ID: {upload_record.financial_record_id}")
                    results.append(("Web文件上传接口", True))
                    print_result(True, "Web文件上传接口测试通过")
                else:
                    results.append(("Web文件上传接口", False))
                    print_result(False, "文件上传记录未保存")
            else:
                print(f"  上传返回: {result}")
                results.append(("Web文件上传接口", False))
                print_result(False, f"文件上传失败: {result.get('error', '未知错误')}")

        except Exception as e:
            results.append(("Web文件上传接口", False))
            print_result(False, f"Web文件上传测试失败: {e}")
            traceback.print_exc()

        print_step(7, "测试文件上传记录查询")
        try:
            from webapp import create_app
            app = create_app()
            client = app.test_client()

            print("\n  查询上传记录列表...")
            response = client.get('/api/uploads')
            data = response.get_json()
            print(f"  总记录数: {data['total']}")

            if data['data']:
                record_id = data['data'][0]['id']
                print(f"\n  查询记录详情 ID={record_id}...")
                response = client.get(f'/api/uploads/{record_id}')
                detail = response.get_json()
                print(f"  文件名: {detail['data']['upload']['file_name']}")
                print(f"  解析状态: {detail['data']['upload']['parse_status']}")
                if detail['data'].get('financial_record'):
                    print(f"  资产负债率: {detail['data']['financial_record'].get('asset_liability_ratio', '-')}")

            results.append(("文件上传记录查询", True))
            print_result(True, "文件上传记录查询测试通过")
        except Exception as e:
            results.append(("文件上传记录查询", False))
            print_result(False, f"文件上传记录查询测试失败: {e}")
            traceback.print_exc()

        print_step(8, "测试DataSyncManager（API未启用模式）")
        try:
            sync_manager = DataSyncManager()

            print("\n  测试CRM同步（API未启用）...")
            result = sync_manager.sync_crm_only()
            print(f"  CRM同步结果: {'已跳过' if result.get('enabled') == False else '执行'}")
            assert result.get('enabled') == False

            print("\n  测试财务系统同步（API未启用）...")
            result = sync_manager.sync_finance_only()
            print(f"  财务同步结果: {'已跳过' if result.get('payments', {}).get('enabled') == False else '执行'}")

            print("\n  测试全量同步（API未启用）...")
            result = sync_manager.sync_all()
            skipped = sum(1 for v in result.values() if v.get('enabled') == False)
            print(f"  全量同步: 跳过 {skipped} 个未启用的API")

            results.append(("DataSyncManager", True))
            print_result(True, "DataSyncManager测试通过（API未启用模式下正确跳过）")
        except Exception as e:
            results.append(("DataSyncManager", False))
            print_result(False, f"DataSyncManager测试失败: {e}")
            traceback.print_exc()

        print("\n" + "="*60)
        print("  测试结果汇总")
        print("="*60)

        passed = sum(1 for _, r in results if r)
        total = len(results)
        print(f"\n总计: {passed}/{total} 项通过")

        for i, (name, result) in enumerate(results, 1):
            status = "✅ 通过" if result else "❌ 失败"
            print(f"  {i}. {name}: {status}")

        print("\n" + "="*60)

        if passed == total:
            print("\n🎉 所有新功能测试通过！")
            return 0
        else:
            print(f"\n⚠️  有 {total - passed} 项测试失败。")
            return 1

    except Exception as e:
        print(f"\n❌ 测试过程中发生严重错误: {e}")
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
