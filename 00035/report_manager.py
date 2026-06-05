import os
import csv
import json
from datetime import datetime, timedelta
from collections import defaultdict
from models import get_connection
from utils import format_datetime, get_time_diff_hours

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    PDF_EXPORT_AVAILABLE = True
except ImportError:
    PDF_EXPORT_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    EXCEL_EXPORT_AVAILABLE = True
except ImportError:
    EXCEL_EXPORT_AVAILABLE = False

class ReportManager:
    def __init__(self, export_dir='reports'):
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)
    
    def generate_weekly_quality_report(self, end_date=None):
        """
        每周生成文档质量报告
        """
        if end_date is None:
            end_date = datetime.now()
        start_date = end_date - timedelta(days=7)
        
        conn = get_connection()
        cursor = conn.cursor()
        
        report_data = {
            'report_period': {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d')
            },
            'review_stats': self._get_review_stats(cursor, start_date, end_date),
            'document_stats': self._get_document_stats(cursor, start_date, end_date),
            'expiry_stats': self._get_expiry_stats(cursor),
            'top_documents': self._get_top_documents(cursor, start_date, end_date)
        }
        
        conn.close()
        
        return {
            'success': True,
            'report_data': report_data,
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    
    def _get_review_stats(self, cursor, start_date, end_date):
        """
        获取审核统计数据
        """
        cursor.execute('''
            SELECT COUNT(*) as total,
                   SUM(CASE WHEN status = 'approved' THEN 1 ELSE 0 END) as approved,
                   SUM(CASE WHEN status = 'rejected' THEN 1 ELSE 0 END) as rejected,
                   SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as pending
            FROM reviews
            WHERE assigned_at BETWEEN ? AND ?
        ''', (start_date.strftime('%Y-%m-%d %H:%M:%S'), 
              end_date.strftime('%Y-%m-%d %H:%M:%S')))
        
        row = cursor.fetchone()
        total = row['total'] or 0
        approved = row['approved'] or 0
        rejected = row['rejected'] or 0
        
        approval_rate = (approved / total * 100) if total > 0 else 0
        
        cursor.execute('''
            SELECT assigned_at, reviewed_at
            FROM reviews
            WHERE status IN ('approved', 'rejected')
              AND assigned_at BETWEEN ? AND ?
              AND reviewed_at IS NOT NULL
        ''', (start_date.strftime('%Y-%m-%d %H:%M:%S'), 
              end_date.strftime('%Y-%m-%d %H:%M:%S')))
        
        reviews = cursor.fetchall()
        avg_duration = 0
        if reviews:
            durations = [get_time_diff_hours(r['assigned_at'], 
                        datetime.strptime(r['reviewed_at'], '%Y-%m-%d %H:%M:%S')) 
                        for r in reviews]
            avg_duration = sum(durations) / len(durations)
        
        return {
            'total_reviews': total,
            'approved': approved,
            'rejected': rejected,
            'pending': row['pending'] or 0,
            'approval_rate': round(approval_rate, 2),
            'avg_review_hours': round(avg_duration, 2)
        }
    
    def _get_document_stats(self, cursor, start_date, end_date):
        """
        获取文档统计数据
        """
        cursor.execute('''
            SELECT COUNT(*) as total_uploaded,
                   SUM(CASE WHEN status = 'published' THEN 1 ELSE 0 END) as published,
                   SUM(view_count) as total_views
            FROM documents
            WHERE created_at BETWEEN ? AND ?
        ''', (start_date.strftime('%Y-%m-%d %H:%M:%S'), 
              end_date.strftime('%Y-%m-%d %H:%M:%S')))
        
        row = cursor.fetchone()
        
        cursor.execute('''
            SELECT doc_type, COUNT(*) as count
            FROM documents
            WHERE created_at BETWEEN ? AND ?
            GROUP BY doc_type
        ''', (start_date.strftime('%Y-%m-%d %H:%M:%S'), 
              end_date.strftime('%Y-%m-%d %H:%M:%S')))
        
        by_type = {row['doc_type']: row['count'] for row in cursor.fetchall()}
        
        return {
            'total_uploaded': row['total_uploaded'] or 0,
            'published': row['published'] or 0,
            'total_views': row['total_views'] or 0,
            'by_type': by_type
        }
    
    def _get_expiry_stats(self, cursor):
        """
        获取过期统计数据
        """
        cursor.execute('SELECT COUNT(*) as total FROM documents WHERE status = \'published\'')
        total_published = cursor.fetchone()['total'] or 0
        
        cursor.execute('SELECT COUNT(*) as expired FROM documents WHERE status = \'expired\'')
        expired = cursor.fetchone()['expired'] or 0
        
        expiry_rate = (expired / (total_published + expired) * 100) if (total_published + expired) > 0 else 0
        
        cursor.execute('''
            SELECT COUNT(*) as expiring_soon
            FROM documents
            WHERE status = 'published'
              AND expiry_date <= datetime('now', '+30 days')
              AND expiry_date > datetime('now')
        ''')
        expiring_soon = cursor.fetchone()['expiring_soon'] or 0
        
        return {
            'total_published': total_published,
            'expired': expired,
            'expiry_rate': round(expiry_rate, 2),
            'expiring_soon': expiring_soon
        }
    
    def _get_top_documents(self, cursor, start_date, end_date, limit=10):
        """
        获取热门文档
        """
        cursor.execute('''
            SELECT id, title, author_name, doc_type, view_count, published_at
            FROM documents
            WHERE status = 'published'
              AND published_at BETWEEN ? AND ?
            ORDER BY view_count DESC
            LIMIT ?
        ''', (start_date.strftime('%Y-%m-%d %H:%M:%S'), 
              end_date.strftime('%Y-%m-%d %H:%M:%S'), limit))
        
        return [dict(row) for row in cursor.fetchall()]
    
    def advanced_query(self, title=None, author=None, status=None, 
                       doc_type=None, start_date=None, end_date=None, 
                       department=None):
        """
        高级查询：按标题、作者、状态、时间段组合查询
        """
        conn = get_connection()
        cursor = conn.cursor()
        
        query = '''
            SELECT d.*, 
                   (SELECT COUNT(*) FROM reviews r WHERE r.document_id = d.id) as review_count,
                   (SELECT GROUP_CONCAT(status) FROM reviews r WHERE r.document_id = d.id) as review_statuses
            FROM documents d
            WHERE 1=1
        '''
        params = []
        
        if title:
            query += ' AND d.title LIKE ?'
            params.append(f'%{title}%')
        
        if author:
            query += ' AND (d.author_name LIKE ? OR d.author_email LIKE ?)'
            params.extend([f'%{author}%', f'%{author}%'])
        
        if status:
            query += ' AND d.status = ?'
            params.append(status)
        
        if doc_type:
            query += ' AND d.doc_type = ?'
            params.append(doc_type)
        
        if start_date:
            query += ' AND d.created_at >= ?'
            params.append(start_date)
        
        if end_date:
            query += ' AND d.created_at <= ?'
            params.append(end_date)
        
        if department:
            query += ' AND d.department = ?'
            params.append(department)
        
        query += ' ORDER BY d.created_at DESC'
        
        cursor.execute(query, params)
        results = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        return {
            'success': True,
            'query_criteria': {
                'title': title, 'author': author, 'status': status,
                'doc_type': doc_type, 'start_date': start_date, 'end_date': end_date,
                'department': department
            },
            'total_count': len(results),
            'results': results
        }
    
    def get_full_lifecycle_record(self, document_id):
        """
        获取文档全生命周期记录
        """
        conn = get_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM documents WHERE id = ?', (document_id,))
        doc = cursor.fetchone()
        
        if not doc:
            conn.close()
            return {'success': False, 'error': '文档不存在'}
        
        cursor.execute('SELECT * FROM reviews WHERE document_id = ? ORDER BY assigned_at', (document_id,))
        reviews = [dict(r) for r in cursor.fetchall()]
        
        cursor.execute('SELECT * FROM document_history WHERE document_id = ? ORDER BY created_at', (document_id,))
        history = [dict(h) for h in cursor.fetchall()]
        
        cursor.execute('SELECT * FROM audit_logs WHERE document_id = ? ORDER BY created_at', (document_id,))
        audit_logs = [dict(a) for a in cursor.fetchall()]
        
        cursor.execute('SELECT * FROM document_readers WHERE document_id = ? ORDER BY read_at', (document_id,))
        readers = [dict(r) for r in cursor.fetchall()]
        
        conn.close()
        
        return {
            'success': True,
            'document': dict(doc),
            'reviews': reviews,
            'version_history': history,
            'audit_logs': audit_logs,
            'read_records': readers
        }
    
    def export_to_csv(self, data, filename=None):
        """
        导出为CSV
        """
        if filename is None:
            filename = f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        file_path = os.path.join(self.export_dir, filename)
        
        if not data:
            return {'success': False, 'error': '没有数据可导出'}
        
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return {
            'success': True,
            'file_path': file_path,
            'filename': filename,
            'record_count': len(data)
        }
    
    def export_to_excel(self, data, filename=None, sheet_name='数据'):
        """
        导出为Excel
        """
        if not EXCEL_EXPORT_AVAILABLE:
            return {'success': False, 'error': '请安装openpyxl库: pip install openpyxl'}
        
        if filename is None:
            filename = f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        file_path = os.path.join(self.export_dir, filename)
        
        if not data:
            return {'success': False, 'error': '没有数据可导出'}
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(row_data.get(key, '')))
                cell.alignment = Alignment(vertical='center')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        wb.save(file_path)
        
        return {
            'success': True,
            'file_path': file_path,
            'filename': filename,
            'record_count': len(data)
        }
    
    def export_weekly_report_to_pdf(self, report_data, filename=None):
        """
        导出周报告为PDF
        """
        if not PDF_EXPORT_AVAILABLE:
            return {'success': False, 'error': '请安装reportlab库: pip install reportlab'}
        
        if filename is None:
            filename = f'weekly_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        file_path = os.path.join(self.export_dir, filename)
        
        doc = SimpleDocTemplate(file_path, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=1
        )
        
        story.append(Paragraph('文档质量周报', title_style))
        story.append(Paragraph(f'报告周期: {report_data["report_period"]["start_date"]} ~ {report_data["report_period"]["end_date"]}', 
                               styles['Normal']))
        story.append(Spacer(1, 20))
        
        review_stats = report_data['review_stats']
        story.append(Paragraph('一、审核统计', styles['Heading2']))
        story.append(Spacer(1, 10))
        
        review_data = [
            ['指标', '数值'],
            ['总审核数', str(review_stats['total_reviews'])],
            ['通过数', str(review_stats['approved'])],
            ['拒绝数', str(review_stats['rejected'])],
            ['待审核数', str(review_stats['pending'])],
            ['审核通过率', f"{review_stats['approval_rate']}%"],
            ['平均审核时长', f"{review_stats['avg_review_hours']} 小时"]
        ]
        
        review_table = Table(review_data, colWidths=[8*cm, 6*cm])
        review_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(review_table)
        story.append(Spacer(1, 20))
        
        expiry_stats = report_data['expiry_stats']
        story.append(Paragraph('二、生命周期统计', styles['Heading2']))
        story.append(Spacer(1, 10))
        
        expiry_data = [
            ['指标', '数值'],
            ['已发布文档总数', str(expiry_stats['total_published'])],
            ['已过期文档数', str(expiry_stats['expired'])],
            ['文档过期率', f"{expiry_stats['expiry_rate']}%"],
            ['即将过期文档(30天内)', str(expiry_stats['expiring_soon'])]
        ]
        
        expiry_table = Table(expiry_data, colWidths=[8*cm, 6*cm])
        expiry_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(expiry_table)
        story.append(Spacer(1, 20))
        
        top_docs = report_data['top_documents']
        if top_docs:
            story.append(Paragraph('三、热门文档TOP10', styles['Heading2']))
            story.append(Spacer(1, 10))
            
            top_data = [['排名', '标题', '作者', '浏览量']]
            for i, doc_item in enumerate(top_docs, 1):
                top_data.append([
                    str(i),
                    doc_item['title'][:20] + '...' if len(doc_item['title']) > 20 else doc_item['title'],
                    doc_item['author_name'],
                    str(doc_item['view_count'])
                ])
            
            top_table = Table(top_data, colWidths=[2*cm, 7*cm, 3*cm, 2*cm])
            top_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(top_table)
        
        doc.build(story)
        
        return {
            'success': True,
            'file_path': file_path,
            'filename': filename
        }
    
    def batch_export_lifecycle_records(self, document_ids, export_format='csv'):
        """
        批量导出生命周期记录
        """
        all_records = []
        
        for doc_id in document_ids:
            record = self.get_full_lifecycle_record(doc_id)
            if record['success']:
                doc = record['document']
                all_records.append({
                    '文档ID': doc['id'],
                    '标题': doc['title'],
                    '类型': doc['doc_type'],
                    '作者': doc['author_name'],
                    '状态': doc['status'],
                    '版本': doc['version'],
                    '创建时间': doc['created_at'],
                    '发布时间': doc['published_at'] or '',
                    '浏览量': doc['view_count'],
                    '审核次数': len(record['reviews']),
                    '历史版本数': len(record['version_history']),
                    '阅读人数': len(record['read_records'])
                })
        
        if export_format == 'csv':
            return self.export_to_csv(all_records)
        elif export_format == 'excel':
            return self.export_to_excel(all_records)
        else:
            return {'success': False, 'error': '不支持的导出格式'}
