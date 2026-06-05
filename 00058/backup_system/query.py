import os
import json
import csv
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Any, Optional, List, Union

from .config import ConfigManager
from .logger import get_logger, log_operation
from .database import DatabaseManager


class QueryManager:
    def __init__(self):
        self._config = ConfigManager()
        self._db = DatabaseManager()
        self._logger = get_logger("query")

    def query_backups(self, system_name: Optional[str] = None,
                     backup_type: Optional[str] = None,
                     start_time: Optional[Union[str, datetime]] = None,
                     end_time: Optional[Union[str, datetime]] = None,
                     status: Optional[str] = None,
                     limit: int = 100,
                     offset: int = 0) -> Dict[str, Any]:
        if isinstance(start_time, str):
            start_time = datetime.fromisoformat(start_time)
        if isinstance(end_time, str):
            end_time = datetime.fromisoformat(end_time)

        backups = self._db.query_backups(
            system_name=system_name,
            backup_type=backup_type,
            start_time=start_time,
            end_time=end_time,
            status=status,
            limit=limit,
            offset=offset
        )

        formatted_backups = [self._format_backup_record(b) for b in backups]

        return {
            'total': len(formatted_backups),
            'limit': limit,
            'offset': offset,
            'records': formatted_backups
        }

    def query_restores(self, system_name: Optional[str] = None,
                      start_time: Optional[Union[str, datetime]] = None,
                      end_time: Optional[Union[str, datetime]] = None,
                      status: Optional[str] = None,
                      limit: int = 100) -> Dict[str, Any]:
        if isinstance(start_time, str):
            start_time = datetime.fromisoformat(start_time)
        if isinstance(end_time, str):
            end_time = datetime.fromisoformat(end_time)

        restores = self._db.query_restores(
            system_name=system_name,
            start_time=start_time,
            end_time=end_time,
            status=status,
            limit=limit
        )

        formatted_restores = [self._format_restore_record(r) for r in restores]

        return {
            'total': len(formatted_restores),
            'records': formatted_restores
        }

    def query_cleanups(self, system_name: Optional[str] = None,
                      start_time: Optional[Union[str, datetime]] = None,
                      end_time: Optional[Union[str, datetime]] = None,
                      limit: int = 100) -> Dict[str, Any]:
        if isinstance(start_time, str):
            start_time = datetime.fromisoformat(start_time)
        if isinstance(end_time, str):
            end_time = datetime.fromisoformat(end_time)

        cleanups = self._db.query_cleanups(
            system_name=system_name,
            start_time=start_time,
            end_time=end_time,
            limit=limit
        )

        return {
            'total': len(cleanups),
            'records': cleanups
        }

    def get_backup_details(self, backup_id: int) -> Optional[Dict[str, Any]]:
        backup = self._db.get_backup_record(backup_id)
        if not backup:
            return None

        file_hashes = self._db.get_file_hashes(backup_id)
        formatted_backup = self._format_backup_record(backup)
        formatted_backup['file_hashes'] = file_hashes

        return formatted_backup

    def get_backup_versions(self, system_name: str, 
                           limit: int = 20) -> List[Dict[str, Any]]:
        backups = self._db.query_backups(
            system_name=system_name,
            status='success',
            limit=limit
        )

        return [
            {
                'backup_id': b['id'],
                'backup_version': b['backup_version'],
                'backup_type': b['backup_type'],
                'start_time': b['start_time'],
                'file_count': b.get('file_count', 0),
                'total_size': b.get('total_size', 0)
            }
            for b in backups
        ]

    def export_backups(self, output_path: str,
                      system_name: Optional[str] = None,
                      backup_type: Optional[str] = None,
                      start_time: Optional[Union[str, datetime]] = None,
                      end_time: Optional[Union[str, datetime]] = None,
                      status: Optional[str] = None,
                      format: str = "csv") -> str:
        if isinstance(start_time, str):
            start_time = datetime.fromisoformat(start_time)
        if isinstance(end_time, str):
            end_time = datetime.fromisoformat(end_time)

        backups = self._db.query_backups(
            system_name=system_name,
            backup_type=backup_type,
            start_time=start_time,
            end_time=end_time,
            status=status,
            limit=10000
        )

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        if format == "csv":
            output_path = self._export_to_csv(backups, output_path)
        elif format == "json":
            output_path = self._export_to_json(backups, output_path)
        elif format == "excel":
            output_path = self._export_to_excel(backups, output_path)
        else:
            raise ValueError(f"Unsupported export format: {format}")

        log_operation("export", system_name or "all", "success",
                     f"Exported {len(backups)} backup records to {format}",
                     {'output_path': output_path, 'count': len(backups)})

        return output_path

    def export_restore_list(self, output_path: str,
                           system_name: Optional[str] = None,
                           format: str = "csv") -> str:
        restores = self._db.query_restores(
            system_name=system_name,
            limit=10000
        )

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        records = [self._format_restore_record(r) for r in restores]

        if format == "csv":
            output_path = self._export_restores_to_csv(records, output_path)
        elif format == "json":
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(records, f, indent=2, ensure_ascii=False, default=str)
        else:
            raise ValueError(f"Unsupported export format: {format}")

        log_operation("export", system_name or "all", "success",
                     f"Exported {len(records)} restore records",
                     {'output_path': output_path})

        return output_path

    def _format_backup_record(self, record: Dict[str, Any]) -> Dict[str, Any]:
        start_time = record['start_time']
        if isinstance(start_time, str):
            start_time = datetime.fromisoformat(start_time)

        end_time = record.get('end_time')
        if end_time and isinstance(end_time, str):
            end_time = datetime.fromisoformat(end_time)

        return {
            'backup_id': record['id'],
            'system_name': record['system_name'],
            'backup_type': record['backup_type'],
            'backup_version': record['backup_version'],
            'source_path': record['source_path'],
            'backup_path': record['backup_path'],
            'file_count': record.get('file_count', 0),
            'total_size': record.get('total_size', 0),
            'total_size_formatted': self._format_size(record.get('total_size', 0)),
            'checksum': record.get('checksum', ''),
            'status': record['status'],
            'retry_count': record.get('retry_count', 0),
            'start_time': start_time,
            'start_time_str': start_time.strftime('%Y-%m-%d %H:%M:%S'),
            'end_time': end_time,
            'end_time_str': end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else '',
            'duration': record.get('duration', 0),
            'error_message': record.get('error_message', '')
        }

    def _format_restore_record(self, record: Dict[str, Any]) -> Dict[str, Any]:
        start_time = record['start_time']
        if isinstance(start_time, str):
            start_time = datetime.fromisoformat(start_time)

        end_time = record.get('end_time')
        if end_time and isinstance(end_time, str):
            end_time = datetime.fromisoformat(end_time)

        return {
            'restore_id': record['id'],
            'system_name': record['system_name'],
            'backup_version': record['backup_version'],
            'backup_id': record.get('backup_id'),
            'restore_path': record['restore_path'],
            'source_hash': record.get('source_hash', ''),
            'restore_hash': record.get('restore_hash', ''),
            'hash_match': record.get('hash_match'),
            'status': record['status'],
            'start_time': start_time,
            'start_time_str': start_time.strftime('%Y-%m-%d %H:%M:%S'),
            'end_time': end_time,
            'end_time_str': end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else '',
            'duration': record.get('duration', 0),
            'error_message': record.get('error_message', ''),
            'diff_report_path': record.get('diff_report_path', '')
        }

    def _export_to_csv(self, backups: List[Dict[str, Any]], output_path: str) -> str:
        if not output_path.endswith('.csv'):
            output_path += '.csv'

        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow([
                '备份ID', '系统名称', '备份类型', '备份版本', '源路径',
                '存储路径', '文件数', '总大小(字节)', '总大小', '校验和',
                '状态', '重试次数', '开始时间', '结束时间', '时长(秒)', '错误信息'
            ])

            for record in backups:
                start_time = record['start_time']
                if isinstance(start_time, str):
                    start_time = datetime.fromisoformat(start_time)
                
                end_time = record.get('end_time')
                if end_time and isinstance(end_time, str):
                    end_time = datetime.fromisoformat(end_time)

                writer.writerow([
                    record['id'],
                    record['system_name'],
                    record['backup_type'],
                    record['backup_version'],
                    record['source_path'],
                    record['backup_path'],
                    record.get('file_count', 0),
                    record.get('total_size', 0),
                    self._format_size(record.get('total_size', 0)),
                    record.get('checksum', ''),
                    record['status'],
                    record.get('retry_count', 0),
                    start_time.strftime('%Y-%m-%d %H:%M:%S'),
                    end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else '',
                    record.get('duration', 0),
                    record.get('error_message', '')
                ])

        self._logger.info(f"Exported {len(backups)} records to CSV: {output_path}")
        return output_path

    def _export_restores_to_csv(self, records: List[Dict[str, Any]], output_path: str) -> str:
        if not output_path.endswith('.csv'):
            output_path += '.csv'

        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow([
                '恢复ID', '系统名称', '备份版本', '备份ID', '恢复路径',
                '源哈希', '恢复哈希', '哈希匹配', '状态',
                '开始时间', '结束时间', '时长(秒)', '错误信息', '差异报告'
            ])

            for r in records:
                writer.writerow([
                    r['restore_id'],
                    r['system_name'],
                    r['backup_version'],
                    r.get('backup_id', ''),
                    r['restore_path'],
                    r.get('source_hash', ''),
                    r.get('restore_hash', ''),
                    r.get('hash_match', ''),
                    r['status'],
                    r['start_time_str'],
                    r['end_time_str'],
                    r.get('duration', 0),
                    r.get('error_message', ''),
                    r.get('diff_report_path', '')
                ])

        return output_path

    def _export_to_json(self, backups: List[Dict[str, Any]], output_path: str) -> str:
        if not output_path.endswith('.json'):
            output_path += '.json'

        formatted = [self._format_backup_record(b) for b in backups]

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(formatted, f, indent=2, ensure_ascii=False, default=str)

        self._logger.info(f"Exported {len(backups)} records to JSON: {output_path}")
        return output_path

    def _export_to_excel(self, backups: List[Dict[str, Any]], output_path: str) -> str:
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "备份清单"

            headers = [
                '备份ID', '系统名称', '备份类型', '备份版本', '源路径',
                '存储路径', '文件数', '总大小', '校验和',
                '状态', '重试次数', '开始时间', '结束时间', '时长(秒)', '错误信息'
            ]

            header_fill = PatternFill(start_color="4299e1", end_color="4299e1", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            for row_idx, record in enumerate(backups, 2):
                start_time = record['start_time']
                if isinstance(start_time, str):
                    start_time = datetime.fromisoformat(start_time)
                
                end_time = record.get('end_time')
                if end_time and isinstance(end_time, str):
                    end_time = datetime.fromisoformat(end_time)

                data = [
                    record['id'],
                    record['system_name'],
                    record['backup_type'],
                    record['backup_version'],
                    record['source_path'],
                    record['backup_path'],
                    record.get('file_count', 0),
                    self._format_size(record.get('total_size', 0)),
                    record.get('checksum', ''),
                    record['status'],
                    record.get('retry_count', 0),
                    start_time.strftime('%Y-%m-%d %H:%M:%S'),
                    end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else '',
                    record.get('duration', 0),
                    record.get('error_message', '')
                ]

                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = center_align
                    cell.border = thin_border

            col_widths = [10, 15, 10, 20, 30, 40, 10, 15, 30, 10, 10, 20, 20, 12, 30]
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            wb.save(output_path)
            self._logger.info(f"Exported {len(backups)} records to Excel: {output_path}")

        except ImportError:
            self._logger.warning("openpyxl not available, falling back to CSV")
            output_path = self._export_to_csv(backups, output_path.replace('.xlsx', ''))

        return output_path

    def get_statistics_summary(self, 
                              start_time: Optional[Union[str, datetime]] = None,
                              end_time: Optional[Union[str, datetime]] = None) -> Dict[str, Any]:
        if isinstance(start_time, str):
            start_time = datetime.fromisoformat(start_time)
        if isinstance(end_time, str):
            end_time = datetime.fromisoformat(end_time)

        if not start_time:
            start_time = datetime.now() - timedelta(days=30)
        if not end_time:
            end_time = datetime.now()

        stats = self._db.get_backup_statistics(start_time, end_time)

        backups = self._db.query_backups(
            start_time=start_time,
            end_time=end_time,
            limit=10000
        )

        system_stats = {}
        for b in backups:
            sys_name = b['system_name']
            if sys_name not in system_stats:
                system_stats[sys_name] = {
                    'total': 0, 'success': 0, 'failed': 0, 'size': 0
                }
            system_stats[sys_name]['total'] += 1
            if b['status'] == 'success':
                system_stats[sys_name]['success'] += 1
                system_stats[sys_name]['size'] += b.get('total_size', 0) or 0
            elif b['status'] == 'failed':
                system_stats[sys_name]['failed'] += 1

        return {
            'period': f"{start_time.strftime('%Y-%m-%d')} to {end_time.strftime('%Y-%m-%d')}",
            'overall': stats,
            'by_system': system_stats
        }

    @staticmethod
    def _format_size(size_bytes: int) -> str:
        import math
        if size_bytes == 0:
            return "0 B"
        size_names = ["B", "KB", "MB", "GB", "TB"]
        i = int(math.floor(math.log(size_bytes, 1024)))
        p = math.pow(1024, i)
        s = round(size_bytes / p, 2)
        return f"{s} {size_names[i]}"
