import os
import json
import math
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple

from .config import ConfigManager
from .logger import get_logger, log_operation
from .database import DatabaseManager
from .notifier import Notifier


class ReportGenerator:
    def __init__(self):
        self._config = ConfigManager()
        self._db = DatabaseManager()
        self._notifier = Notifier()
        self._logger = get_logger("report")

    def generate_weekly_report(self, output_format: str = "pdf") -> Dict[str, Any]:
        self._logger.info("Generating weekly backup report")

        now = datetime.now()
        start_date = now - timedelta(days=7)
        end_date = now

        return self.generate_report(
            report_type="weekly",
            start_date=start_date,
            end_date=end_date,
            output_format=output_format
        )

    def generate_monthly_report(self, output_format: str = "pdf") -> Dict[str, Any]:
        self._logger.info("Generating monthly backup report")

        now = datetime.now()
        start_date = now.replace(day=1) - timedelta(days=1)
        start_date = start_date.replace(day=1)
        end_date = now.replace(day=1) - timedelta(days=1)

        return self.generate_report(
            report_type="monthly",
            start_date=start_date,
            end_date=end_date,
            output_format=output_format
        )

    def generate_report(self, report_type: str, start_date: datetime,
                       end_date: datetime, output_format: str = "pdf",
                       system_name: Optional[str] = None) -> Dict[str, Any]:
        stats = self._db.get_backup_statistics(start_date, end_date)

        period = f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"

        backup_records = self._db.query_backups(
            system_name=system_name,
            start_time=start_date,
            end_time=end_date,
            limit=1000
        )

        restore_records = self._db.query_restores(
            system_name=system_name,
            start_time=start_date,
            end_time=end_date,
            limit=1000
        )

        system_stats = self._get_system_statistics(backup_records)

        report_data = {
            'report_type': report_type,
            'period': period,
            'start_date': start_date,
            'end_date': end_date,
            'generated_at': datetime.now(),
            'statistics': stats,
            'system_stats': system_stats,
            'backup_records': backup_records,
            'restore_records': restore_records
        }

        report_paths = []

        if output_format in ["pdf", "both"]:
            pdf_path = self._generate_pdf_report(report_data)
            report_paths.append(pdf_path)

        if output_format in ["excel", "both"]:
            excel_path = self._generate_excel_report(report_data)
            report_paths.append(excel_path)

        if output_format in ["json", "both"]:
            json_path = self._generate_json_report(report_data)
            report_paths.append(json_path)

        report_id = self._db.create_report_record(
            report_type=report_type,
            report_period=period,
            start_date=start_date,
            end_date=end_date,
            **stats,
            report_path="; ".join(report_paths)
        )

        if report_type == "weekly":
            self._notifier.send_weekly_report(
                report_paths[0] if report_paths else "",
                {**stats, 'period': period}
            )

        log_operation("report", system_name or "all", "success",
                     f"{report_type} report generated",
                     {'report_id': report_id, 'paths': report_paths})

        return {
            'report_id': report_id,
            'report_type': report_type,
            'period': period,
            'statistics': stats,
            'report_paths': report_paths
        }

    def _get_system_statistics(self, backup_records: List[Dict[str, Any]]) -> Dict[str, Any]:
        system_stats = {}

        for record in backup_records:
            sys_name = record['system_name']
            if sys_name not in system_stats:
                system_stats[sys_name] = {
                    'total': 0,
                    'success': 0,
                    'failed': 0,
                    'total_size': 0,
                    'success_size': 0
                }

            system_stats[sys_name]['total'] += 1
            system_stats[sys_name]['total_size'] += record.get('total_size', 0) or 0

            if record['status'] == 'success':
                system_stats[sys_name]['success'] += 1
                system_stats[sys_name]['success_size'] += record.get('total_size', 0) or 0
            elif record['status'] == 'failed':
                system_stats[sys_name]['failed'] += 1

        for sys_name, stats in system_stats.items():
            success_rate = 0.0
            if stats['total'] > 0:
                success_rate = (stats['success'] / stats['total']) * 100
            stats['success_rate'] = round(success_rate, 2)

        return system_stats

    def _generate_pdf_report(self, report_data: Dict[str, Any]) -> str:
        report_dir = self._config.config.log_dir / "reports"
        report_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_path = report_dir / f"backup_report_{report_data['report_type']}_{timestamp}.pdf"

        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, 
                TableStyle, PageBreak
            )
            from reportlab.lib.units import cm

            doc = SimpleDocTemplate(
                str(pdf_path),
                pagesize=landscape(A4),
                rightMargin=2*cm,
                leftMargin=2*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a365d'),
                spaceAfter=20
            )
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#2d3748'),
                spaceAfter=12
            )
            normal_style = styles['Normal']
            normal_style.fontSize = 10

            elements = []

            title = f"备份{report_data['report_type']}报告"
            elements.append(Paragraph(title, title_style))
            elements.append(Paragraph(f"报告周期: {report_data['period']}", normal_style))
            elements.append(Paragraph(f"生成时间: {report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
            elements.append(Spacer(1, 20))

            stats = report_data['statistics']
            elements.append(Paragraph("一、总体统计", heading_style))

            success_rate_color = colors.green if stats['success_rate'] >= 90 else (
                colors.orange if stats['success_rate'] >= 70 else colors.red
            )

            summary_data = [
                ['指标', '数值'],
                ['总备份数', str(stats['total_backups'])],
                ['成功备份', str(stats['successful_backups'])],
                ['失败备份', str(stats['failed_backups'])],
                ['成功率', f"{stats['success_rate']}%"],
                ['总占用空间', self._format_size(stats['total_size'])],
                ['恢复测试数', str(stats['restore_tests'])],
                ['成功恢复数', str(stats['successful_restores'])]
            ]

            summary_table = Table(summary_data, colWidths=[8*cm, 6*cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4299e1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f7fafc')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f7fafc'), colors.white]),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))

            elements.append(Paragraph("二、各系统统计", heading_style))

            system_stats = report_data['system_stats']
            if system_stats:
                sys_data = [['系统名称', '总备份', '成功', '失败', '成功率', '总大小']]
                for sys_name, sys_stat in sorted(system_stats.items()):
                    sys_data.append([
                        sys_name,
                        str(sys_stat['total']),
                        str(sys_stat['success']),
                        str(sys_stat['failed']),
                        f"{sys_stat['success_rate']}%",
                        self._format_size(sys_stat['total_size'])
                    ])

                sys_table = Table(sys_data, colWidths=[5*cm, 2*cm, 2*cm, 2*cm, 2.5*cm, 3.5*cm])
                sys_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#48bb78')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f7fafc'), colors.white]),
                ]))
                elements.append(sys_table)
                elements.append(Spacer(1, 20))

            elements.append(Paragraph("三、备份记录明细", heading_style))

            backup_records = report_data['backup_records']
            if backup_records:
                detail_data = [['系统', '类型', '版本', '状态', '文件数', '大小', '开始时间', '时长(秒)']]
                for record in backup_records[:50]:
                    start_time = record['start_time']
                    if isinstance(start_time, str):
                        start_time = datetime.fromisoformat(start_time)
                    
                    status_color = '✓' if record['status'] == 'success' else (
                        '!' if record['status'] == 'failed' else '-'
                    )
                    
                    detail_data.append([
                        record['system_name'],
                        record['backup_type'],
                        record['backup_version'],
                        status_color,
                        str(record.get('file_count', 0)),
                        self._format_size(record.get('total_size', 0)),
                        start_time.strftime('%m-%d %H:%M'),
                        str(record.get('duration', 0))
                    ])

                detail_table = Table(detail_data, colWidths=[3*cm, 2*cm, 3.5*cm, 1.5*cm, 1.5*cm, 2.5*cm, 3*cm, 2*cm])
                detail_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ed8936')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f7fafc'), colors.white]),
                ]))
                elements.append(detail_table)

            doc.build(elements)
            self._logger.info(f"PDF report generated at {pdf_path}")

        except ImportError as e:
            self._logger.warning(f"reportlab not available, generating text report instead: {e}")
            pdf_path = self._generate_text_report(report_data, ".txt")

        return str(pdf_path)

    def _generate_excel_report(self, report_data: Dict[str, Any]) -> str:
        report_dir = self._config.config.log_dir / "reports"
        report_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = report_dir / f"backup_report_{report_data['report_type']}_{timestamp}.xlsx"

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            header_fill = PatternFill(start_color="4299e1", end_color="4299e1", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            ws1 = wb.active
            ws1.title = "总体统计"

            ws1['A1'] = f"备份{report_data['report_type']}报告"
            ws1['A1'].font = Font(size=20, bold=True, color="1a365d")
            ws1.merge_cells('A1:F1')

            ws1['A2'] = f"报告周期: {report_data['period']}"
            ws1['A2'].font = Font(size=12)
            ws1.merge_cells('A2:F2')

            ws1['A3'] = f"生成时间: {report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}"
            ws1['A3'].font = Font(size=12)
            ws1.merge_cells('A3:F3')

            stats = report_data['statistics']
            summary_headers = ['指标', '数值']
            summary_data = [
                ['总备份数', stats['total_backups']],
                ['成功备份', stats['successful_backups']],
                ['失败备份', stats['failed_backups']],
                ['成功率', f"{stats['success_rate']}%"],
                ['总占用空间', self._format_size(stats['total_size'])],
                ['恢复测试数', stats['restore_tests']],
                ['成功恢复数', stats['successful_restores']]
            ]

            for col_idx, header in enumerate(summary_headers, 1):
                cell = ws1.cell(row=5, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            for row_idx, row_data in enumerate(summary_data, 6):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = center_align
                    cell.border = thin_border

            ws1.column_dimensions['A'].width = 20
            ws1.column_dimensions['B'].width = 20

            ws2 = wb.create_sheet("各系统统计")
            system_headers = ['系统名称', '总备份', '成功', '失败', '成功率', '总大小']
            
            for col_idx, header in enumerate(system_headers, 1):
                cell = ws2.cell(row=1, column=col_idx, value=header)
                cell.fill = PatternFill(start_color="48bb78", end_color="48bb78", fill_type="solid")
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            system_stats = report_data['system_stats']
            row_idx = 2
            for sys_name, sys_stat in sorted(system_stats.items()):
                ws2.cell(row=row_idx, column=1, value=sys_name).alignment = center_align
                ws2.cell(row=row_idx, column=2, value=sys_stat['total']).alignment = center_align
                ws2.cell(row=row_idx, column=3, value=sys_stat['success']).alignment = center_align
                ws2.cell(row=row_idx, column=4, value=sys_stat['failed']).alignment = center_align
                ws2.cell(row=row_idx, column=5, value=f"{sys_stat['success_rate']}%").alignment = center_align
                ws2.cell(row=row_idx, column=6, value=self._format_size(sys_stat['total_size'])).alignment = center_align
                
                for col_idx in range(1, 7):
                    ws2.cell(row=row_idx, column=col_idx).border = thin_border
                row_idx += 1

            for col_idx in range(1, 7):
                ws2.column_dimensions[get_column_letter(col_idx)].width = 15

            ws3 = wb.create_sheet("备份记录")
            backup_headers = ['系统', '类型', '版本', '状态', '文件数', '大小', '开始时间', '时长(秒)', '错误信息']
            
            for col_idx, header in enumerate(backup_headers, 1):
                cell = ws3.cell(row=1, column=col_idx, value=header)
                cell.fill = PatternFill(start_color="ed8936", end_color="ed8936", fill_type="solid")
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            backup_records = report_data['backup_records']
            row_idx = 2
            for record in backup_records:
                start_time = record['start_time']
                if isinstance(start_time, str):
                    start_time = datetime.fromisoformat(start_time)

                ws3.cell(row=row_idx, column=1, value=record['system_name'])
                ws3.cell(row=row_idx, column=2, value=record['backup_type'])
                ws3.cell(row=row_idx, column=3, value=record['backup_version'])
                ws3.cell(row=row_idx, column=4, value=record['status'])
                ws3.cell(row=row_idx, column=5, value=record.get('file_count', 0))
                ws3.cell(row=row_idx, column=6, value=self._format_size(record.get('total_size', 0)))
                ws3.cell(row=row_idx, column=7, value=start_time.strftime('%Y-%m-%d %H:%M:%S'))
                ws3.cell(row=row_idx, column=8, value=record.get('duration', 0))
                ws3.cell(row=row_idx, column=9, value=record.get('error_message', ''))

                for col_idx in range(1, 10):
                    ws3.cell(row=row_idx, column=col_idx).alignment = center_align
                    ws3.cell(row=row_idx, column=col_idx).border = thin_border
                row_idx += 1

            col_widths = [15, 10, 20, 10, 10, 15, 20, 12, 30]
            for col_idx, width in enumerate(col_widths, 1):
                ws3.column_dimensions[get_column_letter(col_idx)].width = width

            wb.save(str(excel_path))
            self._logger.info(f"Excel report generated at {excel_path}")

        except ImportError as e:
            self._logger.warning(f"openpyxl not available, generating CSV report instead: {e}")
            excel_path = self._generate_csv_report(report_data)

        return str(excel_path)

    def _generate_json_report(self, report_data: Dict[str, Any]) -> str:
        report_dir = self._config.config.log_dir / "reports"
        report_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_path = report_dir / f"backup_report_{report_data['report_type']}_{timestamp}.json"

        def serialize(obj):
            if isinstance(obj, datetime):
                return obj.isoformat()
            raise TypeError(f"Type {type(obj)} not serializable")

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False, default=serialize)

        self._logger.info(f"JSON report generated at {json_path}")
        return str(json_path)

    def _generate_text_report(self, report_data: Dict[str, Any], extension: str = ".txt") -> Path:
        report_dir = self._config.config.log_dir / "reports"
        report_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        text_path = report_dir / f"backup_report_{report_data['report_type']}_{timestamp}{extension}"

        stats = report_data['statistics']
        
        content = f"""
{'='*60}
备份{report_data['report_type']}报告
{'='*60}
报告周期: {report_data['period']}
生成时间: {report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}
{'='*60}

一、总体统计
{'-'*60}
总备份数:      {stats['total_backups']}
成功备份:      {stats['successful_backups']}
失败备份:      {stats['failed_backups']}
成功率:        {stats['success_rate']}%
总占用空间:    {self._format_size(stats['total_size'])}
恢复测试数:    {stats['restore_tests']}
成功恢复数:    {stats['successful_restores']}

二、各系统统计
{'-'*60}
"""

        system_stats = report_data['system_stats']
        for sys_name, sys_stat in sorted(system_stats.items()):
            content += f"""
{sys_name}:
  总备份: {sys_stat['total']}, 成功: {sys_stat['success']}, 失败: {sys_stat['failed']}
  成功率: {sys_stat['success_rate']}%, 总大小: {self._format_size(sys_stat['total_size'])}
"""

        content += f"""
三、备份记录明细
{'-'*60}
{'系统':<15} {'类型':<8} {'版本':<18} {'状态':<8} {'文件数':<8} {'大小':<15}
{'-'*60}
"""

        for record in report_data['backup_records'][:50]:
            start_time = record['start_time']
            if isinstance(start_time, str):
                start_time = datetime.fromisoformat(start_time)
            
            content += f"{record['system_name']:<15} {record['backup_type']:<8} {record['backup_version']:<18} "
            content += f"{record['status']:<8} {record.get('file_count', 0):<8} "
            content += f"{self._format_size(record.get('total_size', 0)):<15}\n"

        with open(text_path, 'w', encoding='utf-8') as f:
            f.write(content)

        self._logger.info(f"Text report generated at {text_path}")
        return text_path

    def _generate_csv_report(self, report_data: Dict[str, Any]) -> Path:
        report_dir = self._config.config.log_dir / "reports"
        report_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_path = report_dir / f"backup_report_{report_data['report_type']}_{timestamp}.csv"

        import csv

        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            writer.writerow(['备份报告', report_data['report_type']])
            writer.writerow(['报告周期', report_data['period']])
            writer.writerow([])
            
            stats = report_data['statistics']
            writer.writerow(['指标', '数值'])
            writer.writerow(['总备份数', stats['total_backups']])
            writer.writerow(['成功备份', stats['successful_backups']])
            writer.writerow(['失败备份', stats['failed_backups']])
            writer.writerow(['成功率', f"{stats['success_rate']}%"])
            writer.writerow(['总占用空间', self._format_size(stats['total_size'])])
            writer.writerow([])
            
            writer.writerow(['备份记录明细'])
            writer.writerow(['系统', '类型', '版本', '状态', '文件数', '大小', '开始时间'])
            
            for record in report_data['backup_records']:
                start_time = record['start_time']
                if isinstance(start_time, str):
                    start_time = datetime.fromisoformat(start_time)
                
                writer.writerow([
                    record['system_name'],
                    record['backup_type'],
                    record['backup_version'],
                    record['status'],
                    record.get('file_count', 0),
                    self._format_size(record.get('total_size', 0)),
                    start_time.strftime('%Y-%m-%d %H:%M:%S')
                ])

        self._logger.info(f"CSV report generated at {csv_path}")
        return csv_path

    @staticmethod
    def _format_size(size_bytes: int) -> str:
        if size_bytes == 0:
            return "0 B"
        size_names = ["B", "KB", "MB", "GB", "TB"]
        i = int(math.floor(math.log(size_bytes, 1024)))
        p = math.pow(1024, i)
        s = round(size_bytes / p, 2)
        return f"{s} {size_names[i]}"
